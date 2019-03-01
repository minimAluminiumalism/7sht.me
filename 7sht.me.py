import requests
from requests.exceptions import RequestException
import re
from openpyxl import Workbook
from bs4 import BeautifulSoup


def get_onepage(mainurl):

    response = requests.get(mainurl)
    if response.status_code == 200:
        return response.text

def parse_onepage(html):
    pattern = re.compile('有码</a>\]</em>\s<a href="(.*?)".*?class="s xst">(.*?)\[高清中文字幕\]', re.S)
    items = re.findall(pattern, html)
    return items

#取得磁力连接      
def get_magnet_page(url):
    response = requests.get(url)
    html1 = response.text
    return html1

#下载图片
def get_img(url,UUID):
    response = requests.get(url)
    html = response.text
    pattern = re.compile('class="zoom"\sfile="(.*?)"', re.S)
    items = re.findall(pattern, html)
    for item in items:
        r = requests.get(str(item))
        if r.status_code == 200:
            open(UUID,'wb').write(r.content)


def parse_magnet_page(html1):
    pattern = re.compile('磁力链接.*?<ol><li>(.*?)</ol>', re.S)
    items = re.findall(pattern, html1)
    return items

#创建 excel 文件
def write_to_excel(UUID_LIST, TORRENT_LIST):
	wb = Workbook()
	ws = wb.active
	ws.title = 'AVSheet'

	#创建标题栏
	ws.cell(row=1, column=1, value='影片番号')
	ws.cell(row=1, column=2, value='磁力链接')
	

	index = 3
	for (i, j) in zip(UUID_LIST, TORRENT_LIST):
		ws.cell(row=index, column=1, value=i)
		ws.cell(row=index, column=2, value=j)
		index += 1
	wb.save('new_movies.xlsx')


def main():
	UUID_LIST = []
	TORRENT_LIST = []
	mainurl = 'https://54sadsad.com/forum-103-1.html'
	get_onepage(mainurl)
	html = get_onepage(mainurl)
	parse_onepage(html)

	for item in parse_onepage(html):
		UUID = item[1]
		print(UUID)
		UUID_LIST.append(UUID)
		url = 'https://54sadsad.com/' + item[0]
		html1 = get_magnet_page(url)
		parse_magnet_page(html1)
		get_img(url, UUID)
		for item in parse_magnet_page(html1):
			print(item)
			TORRENT_LIST.append(item)	
	write_to_excel(UUID_LIST, TORRENT_LIST)
       
if __name__ == '__main__':
	main()